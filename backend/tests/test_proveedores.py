"""Tests del módulo de proveedores y valor del dólar."""

from decimal import Decimal
from io import BytesIO

import pytest
from sqlalchemy import text

from app.core.permisos import (
    ROL_CUENTA_MAESTRA,
    ROL_DISTRIBUCION,
    ROL_DUENO,
    ROL_SUPERVISOR,
    Modulo,
    Recurso,
)
from app.models.proveedor import EstadoProveedor, ProveedorDolarHistorial
from app.services import proveedores as servicio
from app.services.roles import NoEncontrado, ReglaDeNegocio


@pytest.fixture
def autor(crear_usuario):
    return crear_usuario("admin", ROL_CUENTA_MAESTRA)


@pytest.fixture
def proveedor(db, autor):
    return servicio.crear_proveedor(
        db, autor, nombre="Distribuidora Norte", dolar_actual=Decimal("1000")
    )


# ============================================================================
# Alta, edición y estado
# ============================================================================


def test_alta_registra_dolar_inicial_en_historial(db, proveedor):
    hist = servicio.historial_dolar(db, proveedor.id)
    assert len(hist) == 1
    assert hist[0].valor_nuevo == Decimal("1000.00")


def test_dolar_no_puede_ser_cero_ni_negativo(db, autor):
    for valor in (Decimal("0"), Decimal("-5")):
        with pytest.raises(ReglaDeNegocio, match="mayor a cero"):
            servicio.crear_proveedor(db, autor, nombre="X", dolar_actual=valor)


def test_baja_es_logica(db, autor, proveedor):
    servicio.cambiar_estado(db, autor, proveedor.id, EstadoProveedor.DESACTIVADO)
    assert proveedor.estado == EstadoProveedor.DESACTIVADO
    # Sigue existiendo en la base.
    assert servicio.obtener_proveedor(db, proveedor.id) is proveedor


def test_desactivado_se_reactiva_libremente(db, crear_usuario, proveedor):
    dist = crear_usuario("dist", ROL_DISTRIBUCION)
    servicio.cambiar_estado(db, dist, proveedor.id, EstadoProveedor.DESACTIVADO)

    # Distribución puede reactivar un desactivado.
    servicio.cambiar_estado(db, dist, proveedor.id, EstadoProveedor.ACTIVO)
    assert proveedor.estado == EstadoProveedor.ACTIVO


def test_inhabilitado_no_lo_reactiva_distribucion(db, crear_usuario, autor, proveedor):
    dist = crear_usuario("dist", ROL_DISTRIBUCION)
    servicio.cambiar_estado(db, autor, proveedor.id, EstadoProveedor.INHABILITADO)

    with pytest.raises(servicio.SinPermiso, match="Cuenta Maestra o Dueño"):
        servicio.cambiar_estado(db, dist, proveedor.id, EstadoProveedor.ACTIVO)


def test_inhabilitado_lo_reactiva_dueno(db, crear_usuario, autor, proveedor):
    dueno = crear_usuario("jefe", ROL_DUENO)
    servicio.cambiar_estado(db, autor, proveedor.id, EstadoProveedor.INHABILITADO)

    servicio.cambiar_estado(db, dueno, proveedor.id, EstadoProveedor.ACTIVO)
    assert proveedor.estado == EstadoProveedor.ACTIVO


def test_no_delete_fisico_en_la_api(client, crear_usuario, roles, dar_permiso, login):
    crear_usuario("admin", ROL_CUENTA_MAESTRA)
    resp = client.delete("/api/v1/proveedores/1", headers=login("admin"))
    # El método no existe en el router.
    assert resp.status_code == 405


# ============================================================================
# Cambio de dólar: individual, masivo, historial
# ============================================================================


def test_cambio_individual_historiza_y_audita(db, autor, proveedor):
    from app.models.auditoria import Auditoria

    servicio.cambiar_dolar(db, autor, proveedor.id, Decimal("1200"))

    assert proveedor.dolar_actual == Decimal("1200.00")
    hist = servicio.historial_dolar(db, proveedor.id)
    assert hist[0].valor_anterior == Decimal("1000.00")
    assert hist[0].valor_nuevo == Decimal("1200.00")

    from sqlalchemy import select

    assert db.execute(
        select(Auditoria).where(Auditoria.accion == "dolar.cambio")
    ).scalars().first() is not None


def test_masivo_porcentaje_por_proveedor(db, autor):
    p1 = servicio.crear_proveedor(db, autor, nombre="A", dolar_actual=Decimal("1000"))
    p2 = servicio.crear_proveedor(db, autor, nombre="B", dolar_actual=Decimal("1500"))

    servicio.cambio_masivo(db, autor, None, "porcentaje", Decimal("10"))

    # El porcentaje se aplica sobre el valor de cada uno.
    assert p1.dolar_actual == Decimal("1100.00")
    assert p2.dolar_actual == Decimal("1650.00")


def test_masivo_valor_fijo(db, autor):
    p1 = servicio.crear_proveedor(db, autor, nombre="A", dolar_actual=Decimal("1000"))
    p2 = servicio.crear_proveedor(db, autor, nombre="B", dolar_actual=Decimal("1500"))

    servicio.cambio_masivo(db, autor, None, "valor", Decimal("2000"))

    assert p1.dolar_actual == Decimal("2000.00")
    assert p2.dolar_actual == Decimal("2000.00")


def test_masivo_genera_un_registro_por_proveedor(db, autor):
    from sqlalchemy import func, select

    p1 = servicio.crear_proveedor(db, autor, nombre="A", dolar_actual=Decimal("1000"))
    p2 = servicio.crear_proveedor(db, autor, nombre="B", dolar_actual=Decimal("1500"))

    servicio.cambio_masivo(db, autor, [p1.id, p2.id], "valor", Decimal("2000"))

    # 2 del alta + 2 del masivo = 4.
    total = db.execute(select(func.count(ProveedorDolarHistorial.id))).scalar_one()
    assert total == 4


def test_masivo_solo_afecta_activos(db, autor):
    activo = servicio.crear_proveedor(db, autor, nombre="A", dolar_actual=Decimal("1000"))
    inactivo = servicio.crear_proveedor(db, autor, nombre="B", dolar_actual=Decimal("1000"))
    servicio.cambiar_estado(db, autor, inactivo.id, EstadoProveedor.DESACTIVADO)

    servicio.cambio_masivo(db, autor, None, "valor", Decimal("2000"))

    assert activo.dolar_actual == Decimal("2000.00")
    assert inactivo.dolar_actual == Decimal("1000.00")  # sin cambios


def test_preview_no_aplica_cambios(db, autor, proveedor):
    servicio.preview_masivo(db, None, "porcentaje", Decimal("50"))
    assert proveedor.dolar_actual == Decimal("1000.00")  # intacto


# ============================================================================
# Importación por Excel
# ============================================================================


def _xlsx(filas: list[tuple]) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["proveedor_id", "dolar_nuevo"])
    for fila in filas:
        ws.append(list(fila))
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_excel_aplica_si_todo_es_valido(db, autor):
    p1 = servicio.crear_proveedor(db, autor, nombre="A", dolar_actual=Decimal("1000"))
    p2 = servicio.crear_proveedor(db, autor, nombre="B", dolar_actual=Decimal("1000"))

    resultado = servicio.importar_dolar(db, autor, _xlsx([(p1.id, 1200), (p2.id, 1300)]))

    assert resultado["aplicados"] == 2
    assert resultado["errores"] == []
    assert p1.dolar_actual == Decimal("1200.00")
    assert p2.dolar_actual == Decimal("1300.00")


def test_excel_es_todo_o_nada(db, autor):
    """Una fila mala aborta la importación completa: criterio de aceptación."""
    p1 = servicio.crear_proveedor(db, autor, nombre="A", dolar_actual=Decimal("1000"))

    resultado = servicio.importar_dolar(
        db, autor, _xlsx([(p1.id, 1200), (999999, 1300)])
    )

    assert resultado["aplicados"] == 0
    assert len(resultado["errores"]) == 1
    assert resultado["errores"][0]["fila"] == 3
    # El proveedor válido NO se modificó.
    assert p1.dolar_actual == Decimal("1000.00")


def test_excel_reporta_valores_invalidos(db, autor):
    p1 = servicio.crear_proveedor(db, autor, nombre="A", dolar_actual=Decimal("1000"))

    resultado = servicio.importar_dolar(
        db, autor, _xlsx([(p1.id, -5), ("no-num", 100), (None, 100)])
    )

    assert resultado["aplicados"] == 0
    assert len(resultado["errores"]) == 3


def test_excel_sin_columnas_correctas(db, autor):
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.append(["id", "valor"])
    buffer = BytesIO()
    wb.save(buffer)

    with pytest.raises(ReglaDeNegocio, match="columnas"):
        servicio.importar_dolar(db, autor, buffer.getvalue())


# ============================================================================
# Endpoints y permisos
# ============================================================================


def test_listado_filtra_por_estado(client, crear_usuario, login, db, autor):
    servicio.crear_proveedor(db, autor, nombre="Activo SA", dolar_actual=Decimal("1000"))
    inact = servicio.crear_proveedor(db, autor, nombre="Baja SA", dolar_actual=Decimal("1000"))
    servicio.cambiar_estado(db, autor, inact.id, EstadoProveedor.DESACTIVADO)

    headers = login("admin")
    resp = client.get("/api/v1/proveedores?estado=activo", headers=headers)

    assert resp.status_code == 200
    assert [p["nombre"] for p in resp.json()] == ["Activo SA"]


def test_sin_permiso_no_lista(client, crear_usuario, login):
    """Vendedor sin permiso sobre proveedores → 403."""
    from app.core.permisos import ROL_VENDEDOR

    crear_usuario("juan", ROL_VENDEDOR)
    resp = client.get("/api/v1/proveedores", headers=login("juan"))
    assert resp.status_code == 403


def test_masivo_requiere_el_recurso(client, crear_usuario, roles, dar_permiso, login):
    """Sin el recurso dolar.cambio_masivo (ni módulo completo) → 403."""
    from app.core.permisos import ROL_VENDEDOR

    crear_usuario("juan", ROL_VENDEDOR)
    dar_permiso(rol_id=roles[ROL_VENDEDOR].id, modulo=Modulo.PROVEEDORES, ver=True)

    resp = client.post(
        "/api/v1/proveedores/dolar/masivo",
        json={"proveedor_ids": None, "modalidad": "valor", "valor": "1000"},
        headers=login("juan"),
    )
    assert resp.status_code == 403


def test_masivo_con_el_recurso_puntual(client, crear_usuario, roles, dar_permiso, login, db, autor):
    """El recurso dolar.cambio_masivo habilita el masivo aunque falte editar general."""
    from app.core.permisos import ROL_VENDEDOR

    servicio.crear_proveedor(db, autor, nombre="A", dolar_actual=Decimal("1000"))
    juan = crear_usuario("juan", ROL_VENDEDOR)
    dar_permiso(
        rol_id=roles[ROL_VENDEDOR].id,
        modulo=Modulo.PROVEEDORES,
        recurso=Recurso.DOLAR_CAMBIO_MASIVO,
        editar=True,
    )

    resp = client.post(
        "/api/v1/proveedores/dolar/masivo",
        json={"proveedor_ids": None, "modalidad": "valor", "valor": "1500"},
        headers=login("juan"),
    )
    assert resp.status_code == 200


# ============================================================================
# NOMBRE (ex razon_social), PAÍS Y PROVINCIA (ex contacto y dirección)
# ============================================================================


def test_el_pais_y_la_provincia_son_opcionales(db, autor):
    """Un proveedor se puede dar de alta sin saber todavía de dónde es."""
    p = servicio.crear_proveedor(
        db, autor, nombre="Sin Datos SA", dolar_actual=Decimal("1000")
    )
    assert (p.pais, p.provincia) == (None, None)


def test_el_pais_y_la_provincia_se_guardan_y_se_editan(db, autor):
    p = servicio.crear_proveedor(
        db, autor, nombre="Distribuidora Norte", dolar_actual=Decimal("1000"),
        pais="Argentina", provincia="Córdoba",
    )
    assert (p.pais, p.provincia) == ("Argentina", "Córdoba")

    servicio.editar_proveedor(db, autor, p.id, pais="Brasil", provincia="São Paulo")
    assert (p.pais, p.provincia) == ("Brasil", "São Paulo")


def test_el_pais_y_la_provincia_viajan_en_la_api(
    client, db, autor, login, dar_permiso, roles
):
    servicio.crear_proveedor(
        db, autor, nombre="Distribuidora Norte", dolar_actual=Decimal("1000"),
        pais="Argentina", provincia="Córdoba",
    )
    db.commit()

    resp = client.get("/api/v1/proveedores", headers=login("admin"))
    assert resp.status_code == 200

    fila = resp.json()[0]
    assert fila["nombre"] == "Distribuidora Norte"
    assert fila["pais"] == "Argentina"
    assert fila["provincia"] == "Córdoba"
    # Los nombres viejos no deben seguir apareciendo en el contrato.
    assert not {"razon_social", "contacto", "direccion"} & set(fila)


def test_el_filtro_de_nombre_sigue_funcionando(db, autor):
    """El rename no puede haberse llevado puesto el filtro del Principio 5."""
    servicio.crear_proveedor(db, autor, nombre="Distribuidora Norte", dolar_actual=Decimal("1000"))
    servicio.crear_proveedor(db, autor, nombre="Mayorista Sur", dolar_actual=Decimal("1000"))

    # ILIKE: insensible a mayúsculas, como el resto de los filtros de texto.
    encontrados = servicio.listar_proveedores(db, nombre="norte")
    assert [p.nombre for p in encontrados] == ["Distribuidora Norte"]


def test_no_quedan_referencias_a_razon_social():
    """
    El rename tiene que ser total: una referencia suelta en un template o
    en el JS falla en silencio (Alpine renderiza vacío, no da error).
    """
    import pathlib

    app = pathlib.Path(__file__).parent.parent / "app"
    con_referencias = [
        str(p.relative_to(app))
        for p in list(app.rglob("*.py")) + list(app.rglob("*.html")) + list(app.rglob("*.js"))
        if "razon_social" in p.read_text()
    ]
    assert not con_referencias, con_referencias


# ============================================================================
# PLANTILLA DESCARGABLE
# ============================================================================


def _hoja_datos(contenido: bytes):
    """Filas de la primera hoja de la plantilla, con el encabezado."""
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(contenido))
    return list(wb["Dolar"].iter_rows(values_only=True))


def test_la_plantilla_se_puede_importar_tal_cual(db, autor, proveedor):
    """
    Es LA garantía del cambio: el formato que se baja y el que se sube son el
    mismo. Si alguna de las dos puntas cambia sin la otra, esto se pone en
    rojo antes de que el usuario descubra que su archivo no entra.
    """
    contenido = servicio.generar_plantilla_dolar(db)

    resultado = servicio.importar_dolar(db, autor, contenido)

    assert resultado["errores"] == []
    # Se leyó bien y no aplicó nada, porque el valor es el mismo que ya
    # tenía: la plantilla se baja con los valores actuales.
    assert resultado["aplicados"] == 0
    assert resultado["sin_cambios"] == 1


def test_la_plantilla_no_trae_proveedores_inactivos(db, autor, proveedor):
    """
    La importación es todo-o-nada y rechaza los inactivos. Con uno solo, el
    archivo bajado sin tocar fallaría entero.
    """
    otro = servicio.crear_proveedor(
        db, autor, nombre="Mayorista Baja", dolar_actual=Decimal("1000")
    )
    servicio.cambiar_estado(db, autor, otro.id, EstadoProveedor.INHABILITADO)
    db.flush()

    nombres = [f[1] for f in _hoja_datos(servicio.generar_plantilla_dolar(db))[1:]]

    assert otro.nombre not in nombres
    assert proveedor.nombre in nombres


def test_la_plantilla_trae_el_valor_actual_y_no_viene_vacia(db, proveedor):
    """
    `dolar_nuevo` va con el valor actual, no en blanco: una fila con
    `proveedor_id` y sin valor NO la saltea el lector —solo saltea las filas
    del todo vacías— y terminaría en "dolar_nuevo inválido o vacío".
    """
    filas = _hoja_datos(servicio.generar_plantilla_dolar(db))

    assert filas[0] == ("proveedor_id", "nombre", "dolar_nuevo")
    fila = filas[1]
    assert fila[0] == proveedor.id
    assert Decimal(str(fila[2])) == proveedor.dolar_actual


def test_la_columna_nombre_no_rompe_la_lectura(db, proveedor):
    """
    Está para que el archivo se pueda leer a ojo. El lector ubica las
    columnas por nombre e ignora las de más, así que no lo molesta — y este
    test lo deja fijado por si alguien cambia esa regla.
    """
    registros = servicio._leer_excel(servicio.generar_plantilla_dolar(db))

    assert len(registros) == 1
    assert int(registros[0]["proveedor_id"]) == proveedor.id


def test_la_plantilla_lleva_una_hoja_de_instrucciones(db, proveedor):
    """
    Las instrucciones van en OTRA hoja: el lector usa `wb.active` y toma todo
    lo que sigue al encabezado como datos, así que un texto arriba de la
    tabla rompería la importación.
    """
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(servicio.generar_plantilla_dolar(db)))

    assert wb.sheetnames[0] == "Dolar"
    assert "Instrucciones" in wb.sheetnames


def test_la_plantilla_por_la_api(client, db, autor, proveedor, login):
    db.commit()

    resp = client.get("/api/v1/proveedores/dolar/plantilla", headers=login("admin"))

    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert "dolar-proveedores-" in resp.headers["content-disposition"]
    # Y lo descargado es un xlsx de verdad, no una respuesta vacía.
    assert _hoja_datos(resp.content)[0][0] == "proveedor_id"


# ============================================================================
# SALTEO DE VALORES IGUALES
# ============================================================================


def test_poner_el_mismo_valor_no_deja_entrada_en_el_historial(db, autor, proveedor):
    """
    El historial existe para registrar cambios. Una entrada "1.400 → 1.400"
    no es un cambio: es ruido en la pantalla donde se va a buscar cuándo se
    movió el precio.
    """
    antes = len(servicio.historial_dolar(db, proveedor.id))

    servicio.cambiar_dolar(db, autor, proveedor.id, proveedor.dolar_actual)

    assert len(servicio.historial_dolar(db, proveedor.id)) == antes


def test_el_salteo_compara_despues_de_redondear(db, autor, proveedor):
    """
    `_validar_dolar` redondea antes de guardar. Comparando sin redondear,
    1400.000 se vería distinto de 1400.00 y el salteo no aplicaría.
    """
    antes = len(servicio.historial_dolar(db, proveedor.id))

    servicio.cambiar_dolar(db, autor, proveedor.id, Decimal("1000.000"))

    assert proveedor.dolar_actual == Decimal("1000.00")
    assert len(servicio.historial_dolar(db, proveedor.id)) == antes


def test_un_valor_distinto_si_se_registra(db, autor, proveedor):
    """El salteo no puede tragarse un cambio real."""
    antes = len(servicio.historial_dolar(db, proveedor.id))

    servicio.cambiar_dolar(db, autor, proveedor.id, Decimal("2000"))

    assert len(servicio.historial_dolar(db, proveedor.id)) == antes + 1
    assert proveedor.dolar_actual == Decimal("2000.00")


def test_el_masivo_saltea_los_que_ya_estaban_en_ese_valor(db, autor, proveedor):
    """
    Poner a todos en el mismo valor no puede dejar una entrada por proveedor:
    solo por los que efectivamente se movieron.
    """
    otro = servicio.crear_proveedor(
        db, autor, nombre="Otro proveedor", dolar_actual=Decimal("1500")
    )
    db.flush()
    antes_p = len(servicio.historial_dolar(db, proveedor.id))
    antes_o = len(servicio.historial_dolar(db, otro.id))

    # `proveedor` ya está en 1000; `otro` está en 1500.
    servicio.cambio_masivo(
        db, autor, proveedor_ids=None, modalidad="valor", valor=Decimal("1000")
    )

    assert len(servicio.historial_dolar(db, proveedor.id)) == antes_p
    assert len(servicio.historial_dolar(db, otro.id)) == antes_o + 1


def test_el_import_informa_cuantas_filas_salteo(db, autor, proveedor):
    """
    No se descartan en silencio: quien sube 10 filas y ve "2 aplicados"
    necesita saber que las otras 8 se leyeron bien, no que se perdieron.
    """
    otro = servicio.crear_proveedor(
        db, autor, nombre="Otro proveedor", dolar_actual=Decimal("1500")
    )
    db.flush()

    from openpyxl import Workbook

    wb = Workbook()
    hoja = wb.active
    hoja.append(["proveedor_id", "dolar_nuevo"])
    hoja.append([proveedor.id, float(proveedor.dolar_actual)])  # igual → saltea
    hoja.append([otro.id, 3000])                                # distinto → aplica
    buffer = BytesIO()
    wb.save(buffer)

    resultado = servicio.importar_dolar(db, autor, buffer.getvalue())

    assert resultado["errores"] == []
    assert resultado["aplicados"] == 1
    assert resultado["sin_cambios"] == 1
