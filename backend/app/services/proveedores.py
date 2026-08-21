"""
Reglas de negocio de proveedores y del valor del dólar.

Todo cambio de `dolar_actual` pasa por acá y deja un registro en
`proveedor_dolar_historial` en la misma transacción, sin importar si vino
de un cambio individual, masivo o por Excel.
"""

from decimal import Decimal, InvalidOperation
from io import BytesIO

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.auditoria import registrar_auditoria, snapshot
from app.core.permisos import ROL_CUENTA_MAESTRA, ROL_DUENO
from app.core.utils import (
    ahora_db,
    normalizar_texto,
    redondear,
    sin_tildes,
    sin_tildes_sql,
)
from app.models.proveedor import (
    EstadoProveedor,
    OrigenCambioDolar,
    Proveedor,
    ProveedorDolarHistorial,
)
from app.models.usuario import Usuario
from app.services.roles import NoEncontrado, ReglaDeNegocio


class SinPermiso(Exception):
    """El autor no puede ejecutar la acción sobre este proveedor (403)."""


# ============================================================================
# LECTURA
# ============================================================================


def obtener_proveedor(db: Session, proveedor_id: int) -> Proveedor:
    proveedor = db.get(Proveedor, proveedor_id)
    if proveedor is None:
        raise NoEncontrado("Proveedor inexistente")
    return proveedor


def listar_proveedores(
    db: Session,
    nombre: str | None = None,
    email: str | None = None,
    telefono: str | None = None,
    provincia: str | None = None,
    pais: str | None = None,
    estado: str | None = None,
    dolar_desde: Decimal | None = None,
    dolar_hasta: Decimal | None = None,
) -> list[Proveedor]:
    """
    Listado con los filtros por defecto del Principio 5, resueltos en el
    backend. Tabla chica: sin paginación, se devuelve todo lo filtrado.
    """
    consulta = select(Proveedor)

    if nombre:
        consulta = consulta.where(Proveedor.nombre.ilike(f"%{nombre}%"))
    if email:
        consulta = consulta.where(Proveedor.email.ilike(f"%{email}%"))
    if telefono:
        consulta = consulta.where(Proveedor.telefono.ilike(f"%{telefono}%"))
    # Las dos caras de la comparación se limpian igual —la columna con
    # `translate()` en SQL y el texto tipeado en Python— para que "cordoba"
    # encuentre "Córdoba", que es como se escribe de apuro en un buscador.
    if provincia:
        consulta = consulta.where(
            sin_tildes_sql(Proveedor.provincia).ilike(f"%{sin_tildes(provincia)}%")
        )
    if pais:
        consulta = consulta.where(
            sin_tildes_sql(Proveedor.pais).ilike(f"%{sin_tildes(pais)}%")
        )
    if estado:
        consulta = consulta.where(Proveedor.estado == estado)
    if dolar_desde is not None:
        consulta = consulta.where(Proveedor.dolar_actual >= dolar_desde)
    if dolar_hasta is not None:
        consulta = consulta.where(Proveedor.dolar_actual <= dolar_hasta)

    consulta = consulta.order_by(Proveedor.nombre)
    return list(db.execute(consulta).scalars().all())


def historial_dolar(db: Session, proveedor_id: int) -> list[ProveedorDolarHistorial]:
    """Historial de cambios del dólar de un proveedor, del más reciente al más viejo."""
    obtener_proveedor(db, proveedor_id)  # 404 si no existe
    return list(
        db.execute(
            select(ProveedorDolarHistorial)
            .where(ProveedorDolarHistorial.proveedor_id == proveedor_id)
            .order_by(ProveedorDolarHistorial.timestamp.desc(), ProveedorDolarHistorial.id.desc())
        )
        .scalars()
        .all()
    )


def _productos_activos(db: Session, proveedor_id: int) -> int:
    """
    Cantidad de productos activos del proveedor.

    La tabla de productos llega en el módulo 04; hasta entonces no existe y
    se asume 0. El chequeo se activa solo cuando la tabla ya está.
    """
    from sqlalchemy import text

    existe = db.execute(
        text("SELECT to_regclass('public.productos')")
    ).scalar()
    if existe is None:
        return 0

    return db.execute(
        text(
            "SELECT count(*) FROM productos "
            "WHERE proveedor_id = :pid AND activo = TRUE"
        ),
        {"pid": proveedor_id},
    ).scalar_one()


# ============================================================================
# ALTA Y EDICIÓN
# ============================================================================


def _validar_dolar(valor: Decimal) -> Decimal:
    """El valor del dólar no puede ser cero ni negativo."""
    if valor is None or valor <= 0:
        raise ReglaDeNegocio("El valor del dólar debe ser mayor a cero")
    return redondear(valor)


def crear_proveedor(
    db: Session,
    autor: Usuario,
    nombre: str,
    dolar_actual: Decimal,
    pais: str | None = None,
    provincia: str | None = None,
    telefono: str | None = None,
    email: str | None = None,
    notas: str | None = None,
    ip_origen: str | None = None,
) -> Proveedor:
    """Alta de proveedor. Registra el dólar inicial también en el historial."""
    nombre_limpio = normalizar_texto(nombre)
    if not nombre_limpio:
        raise ReglaDeNegocio("El nombre es obligatorio")

    dolar = _validar_dolar(dolar_actual)

    proveedor = Proveedor(
        nombre=nombre_limpio,
        pais=normalizar_texto(pais),
        provincia=normalizar_texto(provincia),
        telefono=normalizar_texto(telefono),
        email=normalizar_texto(email),
        notas=normalizar_texto(notas),
        estado=EstadoProveedor.ACTIVO,
        dolar_actual=dolar,
        created_at=ahora_db(),
        updated_at=ahora_db(),
    )
    db.add(proveedor)
    db.flush()

    # El dólar inicial deja rastro en el historial (valor_anterior = nuevo).
    db.add(
        ProveedorDolarHistorial(
            proveedor_id=proveedor.id,
            valor_anterior=dolar,
            valor_nuevo=dolar,
            usuario_id=autor.id,
            origen=OrigenCambioDolar.MANUAL,
            timestamp=ahora_db(),
        )
    )

    registrar_auditoria(
        db,
        usuario_id=autor.id,
        accion="proveedor.crear",
        entidad="proveedores",
        entidad_id=proveedor.id,
        estado_nuevo=proveedor,
        ip_origen=ip_origen,
    )
    db.flush()
    return proveedor


def editar_proveedor(
    db: Session,
    autor: Usuario,
    proveedor_id: int,
    nombre: str | None = None,
    pais: str | None = None,
    provincia: str | None = None,
    telefono: str | None = None,
    email: str | None = None,
    notas: str | None = None,
    ip_origen: str | None = None,
) -> Proveedor:
    """
    Edita los datos del proveedor. NO toca el dólar: ese cambio tiene su
    propio endpoint para que siempre pase por el historial.
    """
    proveedor = obtener_proveedor(db, proveedor_id)
    antes = snapshot(proveedor)

    if nombre is not None:
        nombre_limpio = normalizar_texto(nombre)
        if not nombre_limpio:
            raise ReglaDeNegocio("El nombre es obligatorio")
        proveedor.nombre = nombre_limpio
    if pais is not None:
        proveedor.pais = normalizar_texto(pais)
    if provincia is not None:
        proveedor.provincia = normalizar_texto(provincia)
    if telefono is not None:
        proveedor.telefono = normalizar_texto(telefono)
    if email is not None:
        proveedor.email = normalizar_texto(email)
    if notas is not None:
        proveedor.notas = normalizar_texto(notas)

    proveedor.updated_at = ahora_db()
    db.flush()

    registrar_auditoria(
        db,
        usuario_id=autor.id,
        accion="proveedor.editar",
        entidad="proveedores",
        entidad_id=proveedor.id,
        estado_anterior=antes,
        estado_nuevo=proveedor,
        ip_origen=ip_origen,
    )
    return proveedor


def cambiar_estado(
    db: Session,
    autor: Usuario,
    proveedor_id: int,
    nuevo_estado: EstadoProveedor,
    confirmar_con_productos: bool = False,
    ip_origen: str | None = None,
) -> Proveedor:
    """
    Cambia el estado del proveedor. La baja siempre es lógica; nunca hay
    DELETE físico.

    Reglas:
      - Reactivar un proveedor INHABILITADO requiere Cuenta Maestra o Dueño.
      - No se puede dar de baja un proveedor con productos activos, salvo
        confirmación explícita (`confirmar_con_productos=True`).
    """
    proveedor = obtener_proveedor(db, proveedor_id)
    antes = snapshot(proveedor)

    reactivando = nuevo_estado == EstadoProveedor.ACTIVO
    dando_baja = nuevo_estado in (EstadoProveedor.DESACTIVADO, EstadoProveedor.INHABILITADO)

    # Reactivar algo inhabilitado: solo CM o Dueño.
    if reactivando and proveedor.estado == EstadoProveedor.INHABILITADO:
        rol = autor.rol.nombre if autor.rol else None
        if rol not in (ROL_CUENTA_MAESTRA, ROL_DUENO):
            raise SinPermiso(
                "Reactivar un proveedor inhabilitado requiere Cuenta Maestra o Dueño"
            )

    # Baja con productos activos: exige confirmación.
    if dando_baja:
        activos = _productos_activos(db, proveedor_id)
        if activos and not confirmar_con_productos:
            raise ReglaDeNegocio(
                f"El proveedor tiene {activos} producto(s) activo(s). "
                "Reasignarlos o confirmar la baja explícitamente."
            )

    proveedor.estado = nuevo_estado
    proveedor.updated_at = ahora_db()
    db.flush()

    accion = {
        EstadoProveedor.ACTIVO: "proveedor.reactivar",
        EstadoProveedor.DESACTIVADO: "proveedor.desactivar",
        EstadoProveedor.INHABILITADO: "proveedor.inhabilitar",
    }[nuevo_estado]

    registrar_auditoria(
        db,
        usuario_id=autor.id,
        accion=accion,
        entidad="proveedores",
        entidad_id=proveedor.id,
        estado_anterior=antes,
        estado_nuevo=proveedor,
        ip_origen=ip_origen,
    )
    return proveedor


# ============================================================================
# VALOR DEL DÓLAR
# ============================================================================


def _aplicar_cambio_dolar(
    db: Session,
    proveedor: Proveedor,
    valor_nuevo: Decimal,
    autor: Usuario,
    origen: OrigenCambioDolar,
    ip_origen: str | None,
) -> ProveedorDolarHistorial | None:
    """
    Núcleo compartido por el cambio individual, el masivo y el de Excel:
    valida, actualiza, historiza y audita, todo en la misma transacción.

    Devuelve None cuando el valor nuevo es igual al que ya tenía: en ese
    caso no hace NADA —ni historial, ni auditoría, ni recálculo—.

    El historial existe para registrar cambios, y una entrada
    "1.400 → 1.400" no es un cambio: es ruido que ensucia justo la pantalla
    donde se va a buscar cuándo se movió el precio. Aparecía sobre todo al
    importar el Excel completo habiendo editado dos filas, pero la razón
    vale igual para el cambio individual y el masivo, así que la regla va
    en el punto por el que pasan los tres.
    """
    valor_nuevo = _validar_dolar(valor_nuevo)
    valor_anterior = proveedor.dolar_actual

    # Se compara DESPUÉS de `_validar_dolar`, que redondea: sin eso, 1400.00
    # y 1400.000 se verían distintos y el salteo no aplicaría.
    if valor_nuevo == valor_anterior:
        return None

    registro = ProveedorDolarHistorial(
        proveedor_id=proveedor.id,
        valor_anterior=valor_anterior,
        valor_nuevo=valor_nuevo,
        usuario_id=autor.id,
        origen=origen,
        timestamp=ahora_db(),
    )
    db.add(registro)

    proveedor.dolar_actual = valor_nuevo
    proveedor.updated_at = ahora_db()

    registrar_auditoria(
        db,
        usuario_id=autor.id,
        accion="dolar.cambio",
        entidad="proveedores",
        entidad_id=proveedor.id,
        estado_anterior={"dolar_actual": str(valor_anterior)},
        estado_nuevo={"dolar_actual": str(valor_nuevo), "origen": origen.value},
        ip_origen=ip_origen,
    )
    db.flush()

    # El precio de venta de los productos se deriva de esta cotización y
    # está desnormalizado, así que hay que recalcularlo o la base miente.
    #
    # Va acá y no en los tres endpoints que cambian el dólar (individual,
    # masivo y por Excel) porque los tres desembocan en esta función: es
    # un único lugar que no se puede olvidar. En la misma transacción, así
    # el precio nunca queda a mitad de camino de la cotización.
    #
    # El import es local para no acoplar los módulos al cargarse: el
    # service de productos importa este para validar el alta.
    from app.services.productos import recalcular_precios_de_proveedor

    recalcular_precios_de_proveedor(db, proveedor.id)

    return registro


def cambiar_dolar(
    db: Session,
    autor: Usuario,
    proveedor_id: int,
    valor_nuevo: Decimal,
    ip_origen: str | None = None,
) -> Proveedor:
    """Cambio individual del dólar de un proveedor."""
    proveedor = obtener_proveedor(db, proveedor_id)
    _aplicar_cambio_dolar(db, proveedor, valor_nuevo, autor, OrigenCambioDolar.MANUAL, ip_origen)
    return proveedor


def _calcular_masivo(
    proveedor: Proveedor, modalidad: str, valor: Decimal
) -> Decimal:
    """
    Nuevo valor de un proveedor según la modalidad del cambio masivo.

      - 'valor':      el mismo valor absoluto para todos
      - 'porcentaje': se aplica sobre el dolar_actual de cada uno,
                      redondeado a 2 decimales
    """
    if modalidad == "valor":
        return redondear(valor)
    if modalidad == "porcentaje":
        factor = Decimal(1) + (valor / Decimal(100))
        return redondear(proveedor.dolar_actual * factor)
    raise ReglaDeNegocio(f"Modalidad desconocida: {modalidad!r}")


def preview_masivo(
    db: Session, proveedor_ids: list[int] | None, modalidad: str, valor: Decimal
) -> list[dict]:
    """
    Calcula el resultado del cambio masivo SIN aplicarlo, para el preview.

    `proveedor_ids=None` significa "todos los activos". Solo se consideran
    proveedores en estado ACTIVO.
    """
    proveedores = _proveedores_masivo(db, proveedor_ids)
    return [
        {
            "proveedor_id": p.id,
            "nombre": p.nombre,
            "valor_actual": p.dolar_actual,
            "valor_nuevo": _calcular_masivo(p, modalidad, valor),
        }
        for p in proveedores
    ]


def _proveedores_masivo(db: Session, proveedor_ids: list[int] | None) -> list[Proveedor]:
    """Proveedores activos afectados por un cambio masivo."""
    consulta = select(Proveedor).where(Proveedor.estado == EstadoProveedor.ACTIVO)
    if proveedor_ids:
        consulta = consulta.where(Proveedor.id.in_(proveedor_ids))
    return list(db.execute(consulta.order_by(Proveedor.nombre)).scalars().all())


def cambio_masivo(
    db: Session,
    autor: Usuario,
    proveedor_ids: list[int] | None,
    modalidad: str,
    valor: Decimal,
    ip_origen: str | None = None,
) -> list[dict]:
    """
    Aplica el cambio masivo. Genera un registro de historial y una entrada
    de auditoría por CADA proveedor afectado, no uno global.
    """
    if modalidad not in ("valor", "porcentaje"):
        raise ReglaDeNegocio("Modalidad inválida: usar 'valor' o 'porcentaje'")
    if modalidad == "valor":
        _validar_dolar(valor)  # el valor fijo no puede ser <= 0

    origen = (
        OrigenCambioDolar.MASIVO_VALOR
        if modalidad == "valor"
        else OrigenCambioDolar.MASIVO_PORCENTAJE
    )

    proveedores = _proveedores_masivo(db, proveedor_ids)
    if not proveedores:
        raise ReglaDeNegocio("No hay proveedores activos que coincidan con la selección")

    resultado = []
    for proveedor in proveedores:
        valor_nuevo = _calcular_masivo(proveedor, modalidad, valor)
        _aplicar_cambio_dolar(db, proveedor, valor_nuevo, autor, origen, ip_origen)
        resultado.append(
            {
                "proveedor_id": proveedor.id,
                "nombre": proveedor.nombre,
                "valor_nuevo": valor_nuevo,
            }
        )
    return resultado


# ============================================================================
# IMPORTACIÓN POR EXCEL
# ============================================================================


# Nombres de las columnas que entiende el importador. Se usan tanto para leer
# como para generar la plantilla: si alguna vez cambian, cambian en un solo
# lugar y las dos puntas siguen coincidiendo.
COL_ID = "proveedor_id"
COL_VALOR = "dolar_nuevo"
COL_NOMBRE = "nombre"


def generar_plantilla_dolar(db: Session) -> bytes:
    """
    Excel con todos los proveedores activos y su dólar actual, en el mismo
    formato que espera `importar_dolar`.

    Existe porque sin esto la importación es inusable: hay que armar el
    archivo a mano y averiguar el `id` interno de cada proveedor, que no se
    muestra en ninguna pantalla.

    Tres decisiones que hacen que el archivo se pueda subir tal como se baja:

    - `dolar_nuevo` viene con el valor ACTUAL, no vacío. Una fila con
      `proveedor_id` y sin valor no la saltea el lector —solo saltea las
      filas del todo vacías— y termina en el error "dolar_nuevo inválido o
      vacío".
    - Solo proveedores ACTIVOS. La importación es todo-o-nada y rechaza los
      inactivos, así que incluir uno haría fallar el archivo entero.
    - La columna `nombre` es para poder leer el archivo. El lector ubica las
      columnas por nombre e ignora las de más, así que no lo molesta.
    """
    from openpyxl import Workbook

    proveedores = list(
        db.execute(
            select(Proveedor)
            .where(Proveedor.estado == EstadoProveedor.ACTIVO)
            .order_by(Proveedor.nombre)
        )
        .scalars()
        .all()
    )

    wb = Workbook()
    hoja = wb.active
    hoja.title = "Dolar"
    hoja.append([COL_ID, COL_NOMBRE, COL_VALOR])

    for p in proveedores:
        hoja.append([p.id, p.nombre, float(p.dolar_actual)])

    hoja.column_dimensions["A"].width = 14
    hoja.column_dimensions["B"].width = 38
    hoja.column_dimensions["C"].width = 14

    # Las instrucciones van en OTRA hoja: el lector usa `wb.active` y toma
    # todo lo que sigue al encabezado como datos, así que un texto suelto
    # arriba de la tabla rompería la importación.
    ayuda = wb.create_sheet("Instrucciones")
    for linea in (
        ["Cómo usar esta plantilla"],
        [],
        ["1. Cambiá el valor de la columna dolar_nuevo en los proveedores que quieras."],
        ["2. Subí el archivo en Proveedores → Cambio masivo del dólar."],
        [],
        ["Las filas que dejes sin modificar se saltean: no se registran en el"],
        ["historial ni recalculan precios. No hace falta borrarlas."],
        [],
        ["No cambies los encabezados ni la columna proveedor_id."],
        ["La importación es todo o nada: si una fila falla, no se aplica ningún cambio."],
    ):
        ayuda.append(linea)
    ayuda.column_dimensions["A"].width = 90

    buffer = BytesIO()
    wb.save(buffer)
    wb.close()
    return buffer.getvalue()


def _leer_excel(contenido: bytes) -> list[dict]:
    """
    Lee un .xlsx con columnas `proveedor_id` y `dolar_nuevo`.

    Devuelve una fila por registro con su número de fila (para los mensajes
    de error). No valida contra la base: eso lo hace `importar_dolar`.
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(BytesIO(contenido), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - archivo corrupto o no-xlsx
        raise ReglaDeNegocio("El archivo no es un .xlsx válido") from exc

    hoja = wb.active
    filas = list(hoja.iter_rows(values_only=True))
    wb.close()

    if not filas:
        raise ReglaDeNegocio("El archivo está vacío")

    # Encabezado: se ubican las columnas por nombre, en cualquier orden.
    encabezado = [str(c).strip().lower() if c is not None else "" for c in filas[0]]
    try:
        col_id = encabezado.index(COL_ID)
        col_valor = encabezado.index(COL_VALOR)
    except ValueError as exc:
        raise ReglaDeNegocio(
            f"El Excel debe tener las columnas '{COL_ID}' y '{COL_VALOR}'"
        ) from exc

    registros = []
    for i, fila in enumerate(filas[1:], start=2):
        # Fila totalmente vacía: se ignora.
        if all(c is None or str(c).strip() == "" for c in fila):
            continue
        registros.append(
            {
                "fila": i,
                "proveedor_id": fila[col_id] if col_id < len(fila) else None,
                "dolar_nuevo": fila[col_valor] if col_valor < len(fila) else None,
            }
        )
    return registros


def importar_dolar(
    db: Session, autor: Usuario, contenido: bytes, ip_origen: str | None = None
) -> dict:
    """
    Importa cambios de dólar desde Excel. Es todo-o-nada: si alguna fila
    tiene error, no se aplica ningún cambio y se devuelve la lista completa
    de errores.
    """
    registros = _leer_excel(contenido)
    if not registros:
        raise ReglaDeNegocio("El archivo no tiene filas de datos")

    errores: list[dict] = []
    validos: list[tuple[Proveedor, Decimal]] = []
    vistos: set[int] = set()

    for reg in registros:
        fila = reg["fila"]

        # proveedor_id
        try:
            pid = int(reg["proveedor_id"])
        except (TypeError, ValueError):
            errores.append({"fila": fila, "error": "proveedor_id inválido o vacío"})
            continue

        if pid in vistos:
            errores.append({"fila": fila, "error": f"proveedor_id {pid} repetido en el archivo"})
            continue

        # dolar_nuevo
        try:
            valor = Decimal(str(reg["dolar_nuevo"]))
        except (TypeError, ValueError, InvalidOperation):
            errores.append({"fila": fila, "error": "dolar_nuevo inválido o vacío"})
            continue

        if valor <= 0:
            errores.append({"fila": fila, "error": "dolar_nuevo debe ser mayor a cero"})
            continue

        proveedor = db.get(Proveedor, pid)
        if proveedor is None:
            errores.append({"fila": fila, "error": f"No existe el proveedor {pid}"})
            continue
        if proveedor.estado != EstadoProveedor.ACTIVO:
            errores.append({"fila": fila, "error": f"El proveedor {pid} no está activo"})
            continue

        vistos.add(pid)
        validos.append((proveedor, valor))

    # Todo-o-nada: un solo error aborta la importación completa.
    if errores:
        return {"aplicados": 0, "sin_cambios": 0, "errores": errores}

    # Las filas cuyo valor no cambia se cuentan aparte, no se descartan en
    # silencio: alguien que sube 10 filas y ve "2 aplicados" necesita saber
    # que las otras 8 se leyeron bien y se saltearon a propósito, y no que
    # se perdieron.
    aplicados = 0
    sin_cambios = 0
    for proveedor, valor in validos:
        registro = _aplicar_cambio_dolar(
            db, proveedor, valor, autor, OrigenCambioDolar.IMPORTACION_EXCEL, ip_origen
        )
        if registro is None:
            sin_cambios += 1
        else:
            aplicados += 1

    return {"aplicados": aplicados, "sin_cambios": sin_cambios, "errores": []}
