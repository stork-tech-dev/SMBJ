"""
PDF de una auditoría de inventario: la planilla de lo contado.

Se arma en cada descarga y no se guarda en disco, al revés que el remito. Los
dos motivos:

1. No hay nada que congelar. `auditoria_items` ya guarda `cantidad_sistema`
   —la foto contra la que se comparó— junto con lo contado, así que rearmar
   el documento dentro de un año devuelve exactamente el mismo papel.
2. Guardar una copia sería tener el mismo dato en dos lados (Principio 4),
   con el archivo pudiendo quedar viejo respecto de la base.

El remito es distinto porque es el papel que VIAJÓ con la mercadería: ese sí
es un documento de un momento, y se reimprime tal cual salió.
"""

from io import BytesIO

from sqlalchemy.orm import Session

from app.core.utils import ahora_db
from app.reports.pdf import imprimir
from app.services import configuracion as servicio_configuracion


def generar_pdf_auditoria(db: Session, auditoria) -> bytes:
    """El PDF en bytes, listo para mandar al navegador."""
    return imprimir(
        "reports/auditoria_inventario.html",
        auditoria=auditoria,
        # La letra de la empresa decide la marca, igual que en las pantallas:
        # el mismo sistema atiende a Soleil y a Mallorca.
        letra_empresa=servicio_configuracion.letra_empresa(db),
        # Cuándo se imprimió ESTA copia. Sirve para ordenar dos impresiones
        # de la misma auditoría, que es lo único que las distingue.
        emitido=ahora_db(),
    )


ETIQUETAS_ESTADO = {
    "en_curso": "En curso",
    "pendiente_aprobacion": "Pendiente de aprobación",
    "aprobada": "Aprobada",
    "rechazada": "Rechazada",
}


def generar_xls_auditoria(auditoria) -> bytes:
    """Excel con el detalle de la auditoría, mismo contenido que el PDF."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    hoja = wb.active
    hoja.title = f"Auditoría #{auditoria.id}"

    negrita = Font(bold=True)
    rojo = Font(bold=True, color="CC0000")
    fondo_diff = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # --- Encabezado ---
    estado = auditoria.estado.value if hasattr(auditoria.estado, "value") else auditoria.estado
    hoja.append([f"Auditoría de inventario #{auditoria.id}"])
    hoja["A1"].font = Font(bold=True, size=14)
    hoja.append(["Estado", ETIQUETAS_ESTADO.get(estado, estado)])
    hoja.append(["Ubicación", auditoria.punto_de_venta.nombre])
    hoja.append(["Contó", auditoria.usuario.nombre,
                 auditoria.fecha_inicio.strftime("%d/%m/%Y %H:%M") if auditoria.fecha_inicio else ""])
    aprobador = auditoria.aprobador.nombre if auditoria.aprobador else "—"
    fecha_aprob = auditoria.fecha_aprobacion.strftime("%d/%m/%Y %H:%M") if auditoria.fecha_aprobacion else ""
    hoja.append(["Aprobó", aprobador, fecha_aprob])
    if auditoria.notas:
        hoja.append(["Notas", auditoria.notas])
    hoja.append([])

    # --- Cabecera de tabla ---
    fila_header = ["Código", "Descripción", "Sistema", "Contado", "Diferencia"]
    hoja.append(fila_header)
    for cell in hoja[hoja.max_row]:
        cell.font = negrita

    # --- Items ---
    items = sorted(auditoria.items, key=lambda i: i.variante.codigo_completo)
    for item in items:
        codigo = item.variante.codigo_completo + (item.variante.verificador or "")
        desc = item.variante.producto.descripcion
        if not item.variante.es_base and item.variante.descripcion_sufijo:
            desc += f" · {item.variante.descripcion_sufijo}"
        diff = item.diferencia
        fila = [codigo, desc, item.cantidad_sistema, item.cantidad_contada, diff]
        hoja.append(fila)
        if diff != 0:
            for cell in hoja[hoja.max_row]:
                cell.fill = fondo_diff
            hoja[hoja.max_row][4].font = rojo

    # --- Resumen ---
    hoja.append([])
    con_diff = [i for i in items if i.diferencia != 0]
    hoja.append([f"{len(items)} código(s) contado(s) · {len(con_diff)} con diferencia"])
    if con_diff:
        faltantes = sum(abs(i.diferencia) for i in con_diff if i.diferencia < 0)
        sobrantes = sum(i.diferencia for i in con_diff if i.diferencia > 0)
        hoja.append([f"Faltantes: {faltantes} unidades · Sobrantes: {sobrantes} unidades"])

    # --- Anchos de columna ---
    hoja.column_dimensions["A"].width = 16
    hoja.column_dimensions["B"].width = 40
    hoja.column_dimensions["C"].width = 10
    hoja.column_dimensions["D"].width = 10
    hoja.column_dimensions["E"].width = 12

    # Alinear columnas numéricas a la derecha.
    for row in hoja.iter_rows(min_row=1, max_row=hoja.max_row, min_col=3, max_col=5):
        for cell in row:
            cell.alignment = Alignment(horizontal="right")

    buffer = BytesIO()
    wb.save(buffer)
    wb.close()
    return buffer.getvalue()
